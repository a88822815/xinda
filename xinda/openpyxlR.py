# -*- coding:UTF-8 -*-
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
from itertools import islice

wb = load_workbook(r'C:\Users\lcy\Desktop\mama\re1.xlsx')
ws = wb['Sheet1']
# df0 = pd.DataFrame(ws.values)
data = ws.values
cols = next(data)
cols = next(data)[1:]
data = list(data)
idx = [r[0] for r in data]
data = (islice(r, 1, None) for r in data)
df = pd.DataFrame(data, index=idx, columns=cols)
# df = pd.DataFrame(ws.values)
# df.fillna(0, inplace=True)
df1=df[df['供应商'].isin(['信达'])]
df['票面时间'].iloc[1]==df['票面时间'].iloc[0]
print()

for i in range(0,len(df)):
    # print(i)
    pass