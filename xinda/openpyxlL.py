from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.merge_cells('B2:F4')

# top_left_cell = ws['B2']
# top_left_cell.value = "My Cell"
#
# thin = Side(border_style="thin", color="000000")
# double = Side(border_style="double", color="ff0000")
#
# top_left_cell.border = Border(top=double, left=thin, right=thin, bottom=double)
# top_left_cell.fill = PatternFill("solid", fgColor="DDDDDD")
# top_left_cell.fill = fill = GradientFill(stop=("000000", "FFFFFF"))
# top_left_cell.font  = Font(b=True, color="FF0000")
# top_left_cell.alignment = Alignment(horizontal="center", vertical="center")

border = Border(left=Side(style='thin',color='FF000000'),right=Side(style='thin',color='FF000000'),
                top=Side(style='thin',color='FF000000'),bottom=Side(style='thin',color='FF000000'),
                diagonal=Side(style='medium',color='FF000000'),diagonal_direction=0,
                outline=Side(style='medium',color='FF000000'),vertical=Side(style='medium',color='FF000000'),
                horizontal=Side(style='medium',color='FF000000'))

wb = Workbook()
sht = wb.active
# ws1 = wb.create_sheet("Mysheet", 0)
sht['D1'].value = '入库单'
sht['B2'].value = '供应商:'
sht['D2'].value = '日期:'
sht['F2'].value = '入库单号:'
sht['B3'].value = '品名'
sht['D3'].value = '单位'
sht['E3'].value = '数量'
sht['F3'].value = '价格'
sht['G3'].value = '金额'
sht['B11'].value = '合计'
sht['D11'] = '=SUM(D4:D10)'
sht['E11'] = '=SUM(E4:E10)'
sht['F11'] = '=SUM(F4:F10)'
sht['G11'] = '=SUM(G4:G10)'

font = Font(name=u'宋体', bold = True)
align = Alignment(horizontal='center', vertical='center')
for i in range(3,12):
    a = 'B'+ str(i)
    b = 'C'+ str(i)
    c = a+':'+b
    sht.merge_cells(c)
    sht[a].alignment = align

for i in range(66,72):
    for j in range(3,12):
        a = chr(i) + str(j)
        sht[a].border = border

wb.save('test.xlsx')

